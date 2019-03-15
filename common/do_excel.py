from openpyxl import load_workbook # 读写
from common import contants  # 文件地址

class Cases:  # 第三种方法加的类   存储excel都出来的数据

    def __init__(self):
        self.case_id = None
        self.module = None
        self.title = None
        self.method = None
        self.url = None
        self.data = None
        self.expected = None

class DoExcel:
    '''来获取excel对应表单数据的类 '''
    def  __init__(self, file_name, sheet_name):
        self.file_name = file_name    # 操作文件名
        self.sheet_name = sheet_name   # 操作表单名
        try:
            self.wb = load_workbook(self.file_name)
            self.sheet = self.wb[self.sheet_name]  #这两行在下面的方法中都用到了，所以把他放到初始化函数里面会比较方便
        except Exception as e:
            print("读取excel报错:",e)
            raise e

    def  get_data(self):  # 获取数据

        # 方法三  （重要） 类与对象的思想
        # 读取数据的具体操作
        cases = []  # 存储数据
        for i in range(2, self.sheet.max_row + 1):  # 以行开始读，从第二行开始
            case = Cases()  # 每一行数据存到这个对象里面   对象/实例
            case.case_id = self.sheet.cell(i, 1).value # 存case_id
            case.module = self.sheet.cell(i, 2).value  # 存title
            case.title = self.sheet.cell(i, 3).value  # 存title
            case.method = self.sheet.cell(i, 4).value # 存method
            case.url = self.sheet.cell(i, 5).value  # 存url
            case.data = self.sheet.cell(i, 6).value # 存params
            case.expected = self.sheet.cell(i, 7).value   # 存expected_result
            if type(case.expected) == int:
                case.expected = str(case.expected)
            cases.append(case)

        return cases


    def write_back(self, row, actual_result, test_result):  # 写回数据  行 列 值
        self.sheet.cell(row, 8).value = actual_result  # 写入值
        self.sheet.cell(row, 9).value = test_result

        self.wb.save(self.file_name)  # 保存结果（关闭excel的情况下）


if __name__ == '__main__':
    do_excel = DoExcel(contants.case_dir, 'login')
    test_data = do_excel.get_data()
    print(test_data,test_data[0].__dict__)
    test_data[0].__dict__  # 对象.__dict__可以将一个对象的对应值以字典的形式打印
